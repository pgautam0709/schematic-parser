import csv
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Upload, ParsedRow

router = APIRouter()


@router.get("/export/{upload_id}")
def export_results(upload_id: str, format: str = "csv", db: Session = Depends(get_db)):
    upload = db.get(Upload, upload_id)
    if not upload:
        raise HTTPException(status_code=404, detail="Job not found")
    if upload.status != "done":
        raise HTTPException(status_code=409, detail="Processing not complete")

    rows = (
        db.query(ParsedRow)
        .filter(ParsedRow.upload_id == upload_id)
        .order_by(ParsedRow.sr_number)
        .all()
    )

    safe_name = upload.original_name.replace(" ", "_").replace("(", "").replace(")", "")
    base_name = safe_name.rsplit(".", 1)[0]

    if format == "xlsx":
        return _export_xlsx(rows, base_name)
    return _export_csv(rows, base_name)


def _export_csv(rows, base_name: str) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["SR #", "Page No", "Device", "DT"])
    for row in rows:
        writer.writerow([row.sr_number, row.page_number, row.device or "", row.dt])

    output.seek(0)
    content = "\ufeff" + output.getvalue()  # BOM for Excel compatibility

    return StreamingResponse(
        iter([content.encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{base_name}.csv"'},
    )


def _export_xlsx(rows, base_name: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Device-DT Mapping"

    headers = ["SR #", "Page No", "Device", "DT"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F8F", end_color="2F4F8F", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.sr_number, row.page_number, row.device or "", row.dt])

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{base_name}.xlsx"'},
    )
